"""표 파일(xlsx·csv)을 행 단위 텍스트로.

표형 자료가 RAG에 약한 이유는 임베딩 유사도가 짧은 코드 행과 자연어 질문
사이에서 무너지기 때문이다. 그래서 여기서는 각 행을 "컬럼명: 값 | …" 한 줄로
직렬화한다 - 행이 스스로 컬럼 이름을 들고 다니므로:
  1) RAG 문서 표 조회 MCP(examples_mcp)의 정확 부분일치가 행 단위로 잡고,
     돌려준 창(window) 안에 키·값 이름이 같이 있어 답변 모델이 어떤 값이
     코드인지 문맥으로 읽는다 - 키 컬럼 추측이 필요 없는 이유.
  2) 일반 RAG에도 "컬럼명"이라는 자연어가 행마다 붙어 유사도가 덜 무너진다.

섹션은 시트 이름(csv는 없음): 화면의 청크 좌표와 MCP 폴백의 place가 된다.
section_marker는 무시한다 - 스프레드시트의 행은 계층이 아니다.
"""

import csv
import re
from pathlib import Path

from app.rag.blocks import Block, ParsedDocument
from app.rag.parsers.base import ParseFailure, Parser

EMPTY_MESSAGE = "표에서 읽을 내용이 없습니다. 값이 들어 있는 시트인지 확인해 주세요."


def _row_line(headers: list[str], cells: list[object]) -> str | None:
    """한 행 -> "컬럼: 값 | …". 빈 셀은 건너뛰고, 전부 비면 None."""
    parts = []
    for index, cell in enumerate(cells):
        if cell is None:
            continue
        value = str(cell).strip()
        if not value:
            continue
        header = headers[index].strip() if index < len(headers) and headers[index] else ""
        parts.append(f"{header}: {value}" if header else value)
    return " | ".join(parts) if parts else None


def _blocks_of(rows: list[list[object]], section: str | None) -> list[Block]:
    """첫 번째 비어 있지 않은 행이 머리글이다 - 스프레드시트의 관례이고,
    머리글이 아니었더라도 그 행의 값이 컬럼명 자리에 붙을 뿐 내용은 산다."""
    headers: list[str] | None = None
    blocks: list[Block] = []
    for cells in rows:
        if headers is None:
            if any(c is not None and str(c).strip() for c in cells):
                headers = [str(c).strip() if c is not None else "" for c in cells]
                if section:
                    blocks.append(Block(text=section, block_type="heading", section=section))
            continue
        line = _row_line(headers, cells)
        if line:
            blocks.append(Block(text=line, block_type="paragraph", section=section))
    return blocks


class ExcelParser(Parser):
    """.xlsx - openpyxl read_only 스트리밍, data_only로 수식 대신 계산값."""

    def parse(self, path: str, section_marker: re.Pattern[str] | None = None) -> ParsedDocument:
        from openpyxl import load_workbook  # 무거운 import는 쓰는 순간에

        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            blocks: list[Block] = []
            for sheet in workbook.worksheets:
                rows = [list(row) for row in sheet.iter_rows(values_only=True)]
                blocks.extend(_blocks_of(rows, sheet.title))
        finally:
            workbook.close()
        if not blocks:
            raise ParseFailure(EMPTY_MESSAGE)
        return ParsedDocument(blocks=blocks)


class CsvParser(Parser):
    """.csv - 한국 엑셀이 내보내는 cp949를 utf-8 다음 순서로 시도한다.
    errors=replace로 뭉개면 정확 부분일치 조회가 그 글자를 영원히 못 찾으므로,
    두 인코딩 다 아니면 뭉개는 대신 실패를 말한다."""

    def parse(self, path: str, section_marker: re.Pattern[str] | None = None) -> ParsedDocument:
        raw = Path(path).read_bytes()
        for encoding in ("utf-8-sig", "cp949"):
            try:
                text = raw.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        else:
            raise ParseFailure(
                "CSV의 문자 인코딩을 읽지 못했습니다(UTF-8 또는 CP949로 저장해 주세요)."
            )
        rows = [list(row) for row in csv.reader(text.splitlines())]
        blocks = _blocks_of(rows, None)
        if not blocks:
            raise ParseFailure(EMPTY_MESSAGE)
        return ParsedDocument(blocks=blocks)
